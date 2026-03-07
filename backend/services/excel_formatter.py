"""
Excel Formatter — Styles tracked_jobs.xlsx with color coding,
hyperlinked Apply buttons, and professional formatting.
"""

import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).parent.parent.parent
EXCEL_PATH = ROOT_DIR / "backend" / "tracked_jobs.xlsx"

# ─── Color Palette ────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # Dark slate
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=12)

# Status → row fill color mapping
STATUS_COLORS = {
    "saved":          PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid"),  # Indigo-50
    "github source":  PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid"),  # Green-50
    "shortlisted":    PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid"),  # Amber-50
    "applied":        PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid"),  # Emerald-50
    "interview":      PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid"),  # Blue-50
    "offer":          PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid"),  # Green-50
    "rejected":       PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid"),  # Red-50
    "pending_scrape": PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid"),  # Violet-50
}

# Status → badge font color
STATUS_FONTS = {
    "saved":          Font(name="Segoe UI", color="4338CA", size=11),       # Indigo-700
    "github source":  Font(name="Segoe UI", color="15803D", size=11),       # Green-700
    "shortlisted":    Font(name="Segoe UI", color="B45309", size=11),       # Amber-700
    "applied":        Font(name="Segoe UI", color="047857", size=11, bold=True),  # Emerald-700
    "interview":      Font(name="Segoe UI", color="1D4ED8", size=11, bold=True),  # Blue-700
    "offer":          Font(name="Segoe UI", color="15803D", size=11, bold=True),  # Green-700
    "rejected":       Font(name="Segoe UI", color="B91C1C", size=11),       # Red-700
    "pending_scrape": Font(name="Segoe UI", color="6D28D9", size=11),       # Violet-700
}

DEFAULT_FILL = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
DEFAULT_FONT = Font(name="Segoe UI", size=11, color="374151")
LINK_FONT    = Font(name="Segoe UI", size=11, color="2563EB", underline="single")
APPLY_FONT   = Font(name="Segoe UI", size=11, color="FFFFFF", bold=True)
APPLY_FILL   = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")  # Indigo-600

THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
WRAP   = Alignment(horizontal="left", vertical="center", wrap_text=True)


from backend.services.db_tracker import get_jobs

def sync_db_to_excel(output_excel: Path = EXCEL_PATH):
    jobs = get_jobs()
    if not jobs:
        return
        
    df = pd.DataFrame(jobs)
    df_mapped = pd.DataFrame()
    df_mapped["Title"] = df.get("title", "")
    df_mapped["Company"] = df.get("company", "")
    df_mapped["URL"] = df.get("apply_link", "")
    df_mapped["Status"] = df.get("status", "")
    df_mapped["Score"] = df.get("score", "")
    df_mapped["Reason"] = df.get("reason", "")
    df_mapped["Date Added"] = df.get("found_at", "")
    
    df_mapped.to_excel(output_excel, index=False, engine="openpyxl")
    format_excel(output_excel)

def format_excel(path: Path = EXCEL_PATH) -> dict:
    """Read, format, and overwrite the tracked_jobs.xlsx with rich styling.

    Returns {"status": "success", "rows": int} or raises.
    """
    if not path.exists():
        return {"status": "error", "message": "Excel file not found."}

    # ── Load data with pandas, then enhance with openpyxl ──
    df = pd.read_excel(path)

    # Ensure expected columns exist
    expected = ["Title", "Company", "URL", "Status", "Date Added"]
    for col in expected:
        if col not in df.columns:
            df[col] = ""

    # Add "Apply" column if missing
    if "Apply" not in df.columns:
        df["Apply"] = "➜ Apply"

    # Reorder columns
    final_cols = ["Title", "Company", "URL", "Status", "Score", "Reason", "Date Added", "Apply"]
    df = df[[c for c in final_cols if c in df.columns]]

    # Write to Excel first (clean slate)
    df.to_excel(path, index=False, engine="openpyxl")

    # ── Now open with openpyxl for styling ──
    wb = load_workbook(path)
    ws = wb.active
    ws.title = "Job Tracker"

    num_rows = ws.max_row
    num_cols = ws.max_column

    # ── Column widths ──
    col_widths = {
        "Title": 42,
        "Company": 24,
        "URL": 50,
        "Status": 18,
        "Score": 10,
        "Reason": 55,
        "Date Added": 22,
        "Apply": 14,
    }
    for col_idx in range(1, num_cols + 1):
        header_val = ws.cell(row=1, column=col_idx).value
        width = col_widths.get(header_val, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Row height ──
    ws.row_dimensions[1].height = 32
    for row in range(2, num_rows + 1):
        ws.row_dimensions[row].height = 28

    # ── Style header row ──
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ── Find column indices ──
    col_map = {}
    for col_idx in range(1, num_cols + 1):
        col_map[ws.cell(row=1, column=col_idx).value] = col_idx

    url_col = col_map.get("URL")
    status_col = col_map.get("Status")
    apply_col = col_map.get("Apply")

    # ── Style data rows ──
    for row in range(2, num_rows + 1):
        status_val = ""
        if status_col:
            status_cell = ws.cell(row=row, column=status_col)
            status_val = str(status_cell.value or "").strip().lower()

        row_fill = STATUS_COLORS.get(status_val, DEFAULT_FILL)
        status_font = STATUS_FONTS.get(status_val, DEFAULT_FONT)

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = THIN_BORDER
            cell.alignment = CENTER if col_idx != col_map.get("Title", -1) else WRAP

            # Apply row fill based on status
            cell.fill = row_fill
            cell.font = DEFAULT_FONT

            # Status column gets special font color
            if col_idx == status_col:
                cell.font = status_font
                cell.alignment = CENTER

            # URL column → clickable hyperlink
            if col_idx == url_col and cell.value:
                url_str = str(cell.value)
                if url_str.startswith("http"):
                    cell.hyperlink = url_str
                    cell.font = LINK_FONT
                    # Shorten display text
                    if len(url_str) > 60:
                        cell.value = url_str[:57] + "..."

            # Apply column → styled button with hyperlink
            if col_idx == apply_col:
                url_cell = ws.cell(row=row, column=url_col) if url_col else None
                url_value = str(url_cell.value or "") if url_cell else ""
                # Try to get the full URL (it might have been shortened)
                if url_cell and url_cell.hyperlink:
                    url_value = url_cell.hyperlink.target or url_value

                cell.value = "➜ Apply"
                cell.font = APPLY_FONT
                cell.fill = APPLY_FILL
                cell.alignment = CENTER

                if url_value.startswith("http"):
                    cell.hyperlink = url_value

    # ── Freeze header row ──
    ws.freeze_panes = "A2"

    # ── Auto-filter ──
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    return {"status": "success", "rows": num_rows - 1, "path": str(path)}
